# papulic 2018
# encoding=utf8
import os
import shutil
import openpyxl
import time
import sys
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import sys
reload(sys)
sys.setdefaultencoding('utf8')

# #####################################################################################################################
# #####################################################################################################################
# #####################################################################################################################

# cuvanje svih rezultata (stanje u svim magacinima i ostali podaci se cuvaju u tekstualnim fajlovima)
CUVAJ_SVE = True

# pauza je u sekundama!!!
PAUZA_IZMEDJU_PRETRAGA = 10

# korisnik za gazelu
korisnik_gazela = "fiatino"
lozinka_gazela = "fiatino123"

# korisnik za wint
korisnik_wint = "autotigra"
lozinka_wint = "tigra2167300"

# korisnik za rv trade
korisnik_rvtrade = "TESTRV1"
lozinka_rvtrade = "RVadmin021"

# #####################################################################################################################
# #####################################################################################################################
# #####################################################################################################################

param_error = """
Pogresan broj parametara..
Kako koristiti skriptu: rv_trade.py <textualni fajl sa spiskom artikala za pretragu> <sajt za pretragu 1> <sajt za pretragu 2> ...
Moguci sajtovi: 'gazela', 'wint', 'rvtrade'
"""
STRING_PAUZA = str(PAUZA_IZMEDJU_PRETRAGA)
MOGUCI_SAJTOVI = ['gazela', 'wint', 'rvtrade']

# sajtovi
gazela = False
wint = False
rvtrade = False
# Artikli za pretragu
artikli_za_pretragu = []

# Provera ispravnosti parametara
if not len(sys.argv) > 2:
    sys.exit(param_error)
else:
    try:
        lista = open(sys.argv[1]).readlines()
    except IOError:
        sys.exit("Ne postoji fajl: '{0}'!".format(sys.argv[1]))
    for artikal in lista:
        if artikal != "\n":
            artikal = artikal.replace("\n", "")
            if artikal not in artikli_za_pretragu:
                artikli_za_pretragu.append(artikal)
    sajtovi = sys.argv[2:]
    for sajt in sajtovi:
        if sajt not in MOGUCI_SAJTOVI:
            sys.exit("Pogresan sajt: '{0}'!".format(sajt))
        else:
            if sajt == "gazela":
                gazela = True
            elif sajt == "wint":
                wint = True
            elif sajt == "rvtrade":
                rvtrade = True

# excel
XLSX_FILE_TEMP = "rezultati_temp.xlsx"
XLSX_FILE = "rezultati.xlsx"

# ____________________________________________________________________________________________________
if os.path.exists(XLSX_FILE_TEMP):
    book = openpyxl.load_workbook(XLSX_FILE_TEMP)
else:
    wb = openpyxl.Workbook()
    for i in MOGUCI_SAJTOVI:
        wb.create_sheet(i)
    wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
    wb.save(XLSX_FILE_TEMP)
    book = openpyxl.load_workbook(XLSX_FILE_TEMP)


if gazela:
    # get the path of ChromeDriverServer
    dir = os.path.dirname(__file__)
    chrome_driver_path = "chromedriver.exe"

    # create a new Chrome session
    driver = webdriver.Chrome(chrome_driver_path)
    driver.implicitly_wait(30)
    # mora full screen , u suprotnom nece naci pretragu !!
    driver.maximize_window()


    driver.get("http://gazela.co.rs/")
    username = driver.find_element_by_id("txtUserName")
    password = driver.find_element_by_id("txtPass")

    username.send_keys(korisnik_gazela)
    password.send_keys(lozinka_gazela)

    driver.find_element_by_id("Button1").click()
    for artikal in artikli_za_pretragu:
        print "Pretrazujem artikal '{0}'...\n".format(artikal)
        time.sleep(2)

        # pretraga
        nasao_pretragu = False
        while not nasao_pretragu:
            try:
                driver.find_element_by_xpath('//a[@href="TopMotive.aspx"]').click()
                nasao_pretragu = True
            except:
                pass

        time.sleep(2)

        driver.switch_to.frame(driver.find_element_by_id("Main_iframe1"))

        searchbox = driver.find_element_by_id("home_txt_art_direkt")

        # unesi artikal za pretragu
        searchbox.send_keys(artikal)
        time.sleep(1)

        driver.find_element_by_id("home_imgBtn_art_direkt").click()

        time.sleep(3)
        html = driver.page_source
        # ucitavanje cele stranice
        soup = BeautifulSoup(html, "lxml")
        uspesna_pretraga = True

        tabela = soup.find('table', id="main_artikel_panel_maintable")
        if tabela != None:
            brendovi = []
            for row in tabela.find_all('tr'):
                if row.has_attr('class'):
                    if "main_artikel_panel_tr_einspeiser" in row.attrs['class']:
                        columns = row.find_all('td')
                        for col in columns:
                            if col.has_attr('colspan'):
                                brend = col.find('span').text
                                brendovi.append([brend])
                    elif "main_artikel_panel_tr_artikel" in row.attrs['class']:
                        if "artikel1" in row.attrs['row_type']:
                            brendovi[-1].append({})
                            status_i_cena = row.find_all('td')
                            for col in status_i_cena:
                                if col.has_attr('class'):
                                    if 'tc_number' in col.attrs['class']:
                                        sifre = col.find_all('a')
                                        for a in sifre:
                                            if a.has_attr('title'):
                                                if a.attrs['title'].startswith("Sifra dobavljacevog artikla"):
                                                    sifra_dobavljacevog_artikla = a.contents[0].text
                                                    brendovi[-1][-1]["sifra dobavljacevog artikla"] = sifra_dobavljacevog_artikla
                                                elif a.attrs['title'].startswith("Kataloski broj"):
                                                    kataloski_broj_proizvodjaca = a.contents[0].text
                                                    brendovi[-1][-1]["kataloski broj proizvodjaca"] = kataloski_broj_proizvodjaca

            cene = driver.find_elements_by_tag_name('img')
            # sve_cene = []
            for cena in cene:
                title = cena.get_attribute('title')
                if title.startswith('Pitanje o trenutnim kolicinama/ceni'):
                    # sve_cene.append(cena)
                    if cena.is_displayed():
                        modal_ok = False
                        while not modal_ok:
                            driver.execute_script("arguments[0].click();", cena)
                            time.sleep(2)
                            driver.switch_to.frame(driver.find_elements_by_class_name("cboxIframe")[0])
                            time.sleep(1)
                            master = driver.find_elements_by_id("masterpane")[0]
                            html_text = master.text
                            html_cena = master.text.split('\n')


                            if not html_text.startswith("Prilikom obrade pitanja je nastala greska"):
                                modal_ok = True
                            if not modal_ok:
                                driver.switch_to.default_content()
                                time.sleep(1)
                                driver.switch_to.frame(driver.find_element_by_id("Main_iframe1"))
                                time.sleep(1)
                                element = driver.find_element_by_id('cboxClose')
                                driver.execute_script("arguments[0].click();", element)
                                time.sleep(1)

                        if len(html_cena) > 11:
                            artikal_pretraga = html_cena[0]
                            stanje = html_cena[3]
                            if stanje.startswith("Dostupno"):
                                stanje = "IMA"
                            elif stanje.startswith("Nije dostupno"):
                                stanje = "NEMA"
                            else:
                                stanje = "NEPOZNATO"
                            for index, i in enumerate(html_cena):
                                if i.startswith('VP Cena') and not i.startswith('VP cena sa popustom'):
                                    vp_cena = i
                                elif i.startswith('Magacin'):
                                    lokacije = html_cena[index + 1:]
                            # vp_cena = html_cena[11]
                            # lokacije = html_cena[17:]
                            lokacije_ima = []
                            for location in lokacije:
                                loc, state = location.split()
                                if state == "Ima":
                                    lokacije_ima.append(loc)
                            for brend in brendovi:
                                for num, a in enumerate(brend):
                                    if num > 0:
                                        try:
                                            if a['sifra dobavljacevog artikla'] == artikal_pretraga:
                                                a["vp cena"] = vp_cena
                                                a["na stanju"] = stanje
                                                a["lokacije"] = lokacije_ima
                                                print artikal_pretraga
                                                # if CUVAJ_SVE:
                                                #     if not os.path.exists("rezultati"):
                                                #         os.makedirs("rezultati")
                                                #     ime_fajla = artikal + "-" + brend[0] + "-" + a["kataloski broj proizvodjaca"]
                                                #     for ch in ['&', '#', '/', '\\', '<', '>', '"', '?', "'", ':', ';', '|', '*', '!', ' ']:
                                                #         if ch in ime_fajla:
                                                #             ime_fajla = ime_fajla.replace(ch, "")
                                                #     temp = open("rezultati\\" + ime_fajla + ".txt", 'w')
                                                #     temp.write(html_text)
                                                #     temp.close()
                                        except KeyError:
                                            pass

                        driver.switch_to.default_content()
                        time.sleep(1)
                        driver.switch_to.frame(driver.find_element_by_id("Main_iframe1"))
                        time.sleep(1)
                        element = driver.find_element_by_id('cboxClose')
                        driver.execute_script("arguments[0].click();", element)
                        time.sleep(1)

        worksheet = book['gazela']
        row_idx = worksheet.max_row + 2
        worksheet.cell(row=row_idx, column=1).value = artikal
        worksheet.cell(row=row_idx + 1, column=2).value = "HEAD1"
        worksheet.cell(row=row_idx + 1, column=3).value = "HEAD2"
        worksheet.cell(row=row_idx + 1, column=4).value = "HEAD3"
        worksheet.cell(row=row_idx + 1, column=5).value = "HEAD4"
        row_idx += 2
        if tabela != None:
            for brend in brendovi:
                for num, i in enumerate(brend):
                    if num > 0:
                        worksheet.cell(row=row_idx, column=2).value = brend[0]
                        if 'kataloski broj proizvodjaca' in i:
                            worksheet.cell(row=row_idx, column=3).value = i['kataloski broj proizvodjaca']
                        else:
                            worksheet.cell(row=row_idx, column=3).value = '---'
                        if 'vp cena' in i:
                            try:
                                if i['vp cena'].startswith("VP Cena(Rasprodaja do isteka zaliha) "):
                                    vp_cena = i['vp cena'].split('VP Cena(Rasprodaja do isteka zaliha) ')[1].split(' RSD')[0]
                                else:
                                        vp_cena = i['vp cena'].split('VP Cena ')[1].split(' RSD')[0]
                                if vp_cena[0].isdigit():
                                    vp_cena = float(vp_cena)
                            except IndexError:
                                vp_cena = i['vp cena']
                            worksheet.cell(row=row_idx, column=4).value = vp_cena
                        else:
                            worksheet.cell(row=row_idx, column=4).value = '---'
                        if 'na stanju' in i:
                            worksheet.cell(row=row_idx, column=5).value = i['na stanju']
                            if CUVAJ_SVE:
                                if 'lokacije' in i:
                                    for pos, lokacija in enumerate(i['lokacije']):
                                        worksheet.cell(row=row_idx, column=6 + pos).value = lokacija
                        else:
                            worksheet.cell(row=row_idx, column=5).value = '---'
                        row_idx += 1
        else:
            worksheet.cell(row=row_idx, column=2).value = "Ne postoji ovaj artikal"
            row_idx += 1
        book.save(XLSX_FILE_TEMP)
        if os.path.exists(XLSX_FILE):
            os.remove(XLSX_FILE)
        shutil.copyfile(XLSX_FILE_TEMP, XLSX_FILE)
        driver.switch_to.default_content()
        if artikal != artikli_za_pretragu[-1]:
            print "\nCekam {sekunde} sekundi".format(sekunde=STRING_PAUZA)
            time.sleep(PAUZA_IZMEDJU_PRETRAGA)
    driver.quit()
    print "\nUspesna pretraga!"

if wint:
    # get the path of ChromeDriverServer
    dir = os.path.dirname(__file__)
    chrome_driver_path = "chromedriver.exe"

    # create a new Chrome session
    driver = webdriver.Chrome(chrome_driver_path)
    driver.implicitly_wait(30)
    # mora full screen , u suprotnom nece naci pretragu !!
    driver.maximize_window()

    driver.get("http://109.245.231.209/WagenWebShop/")
    username = driver.find_element_by_id("username")
    password = driver.find_element_by_id("password")

    username.send_keys(korisnik_wint)
    password.send_keys(lozinka_wint)

    driver.find_element_by_xpath("/html/body/div/div[2]/div/div[2]/form/div/div[3]/button").click()

    driver.find_element_by_xpath("/html/body/div[5]/div/img").click()



    time.sleep(2)



    for _artikal_ in artikli_za_pretragu:
        print "Pretrazujem artikal '{0}'...\n".format(_artikal_)

        nasao_pretragu = False
        trys = 0
        while not nasao_pretragu:
            try:
                driver.find_element_by_xpath('//a[@href="/WagenWebShop/Shop/QuickShop"]').click()
                nasao_pretragu = True
            except:
                trys += 1
                if trys > 10:
                    close_no_item = driver.find_elements_by_tag_name('button')
                    for button in close_no_item:
                        if "ui-corner-all" in button.get_attribute("class"):
                            if button.text == "Ok":
                                button.click()
                pass

        time.sleep(2)

        searchbox = driver.find_element_by_id("tbProductNo")

        # unesi artikal za pretragu
        searchbox.send_keys(_artikal_)
        time.sleep(0.5)

        searchbox.send_keys(Keys.ENTER)

        time.sleep(3)

        search_tr = driver.find_element_by_id("trResult")

        uspesna_pretraga = True

        # hover sve elemente da se ucitaju podaci o stanjima
        stanja = driver.find_elements_by_tag_name('a')
        for s in stanja:
            klasa = s.get_attribute("class")
            if klasa == "tt":
                element_to_hover_over = s

                hover = ActionChains(driver).move_to_element(element_to_hover_over)
                hover.perform()
                time.sleep(1)
                element_to_hover_over = search_tr

                hover = ActionChains(driver).move_to_element(element_to_hover_over)
                hover.perform()
                time.sleep(1)
        html = driver.execute_script("return document.getElementsByTagName('html')[0].innerHTML")
        time.sleep(1)
        soup = BeautifulSoup(html, "lxml")

        tabela = soup.find('table', id="productsTable")

        artikli_pretrage = []
        head = tabela.find('thead')# .text.split('\n')
        body = tabela.find('tbody')
        svi_artikli = body.find_all('tr')
        for a in svi_artikli:
            if a.has_attr("id"):
                if a.attrs['id'].startswith("trResult"):
                    continue
            podaci = a.find_all('td')
            if len(podaci) < 3:
                continue
            artikal_recnik = {}
            sifra = podaci[0].text
            OEbroj = podaci[1].text
            opis = podaci[2].text
            brend = podaci[4].text
            str = podaci[5].text.replace(".", "").replace(",", ".")
            cena = round(float(str) * 1.111, 2)

            artikal_recnik['sifra'] = sifra
            artikal_recnik['OEbroj'] = OEbroj
            artikal_recnik['opis'] = opis
            artikal_recnik['brend'] = brend
            artikal_recnik['cena'] = cena
            stanje = podaci[6].find('table')
            nova_stanja = []
            if stanje is not None:
                stanje = stanje.text.split("\n")

                for ind, k in enumerate(stanje):
                    if not k == "" and not k == "Lokacija" and not k == "Stanje":
                        nova_stanja.append(k)
            else:
                nova_stanja.append("GRESKA")
            artikal_recnik['stanje'] = nova_stanja
            artikli_pretrage.append(artikal_recnik)
        if artikli_pretrage != []:
            worksheet = book['wint']
            row_idx = worksheet.max_row + 2
            worksheet.cell(row=row_idx, column=1).value = _artikal_
            worksheet.cell(row=row_idx + 1, column=2).value = "SIFRA"
            worksheet.cell(row=row_idx + 1, column=3).value = "OE_BROJ"
            worksheet.cell(row=row_idx + 1, column=4).value = "OPIS"
            worksheet.cell(row=row_idx + 1, column=5).value = "BREND"
            worksheet.cell(row=row_idx + 1, column=6).value = "CENA"
            worksheet.cell(row=row_idx + 1, column=7).value = "STANJE"
            row_idx += 2
            for artikal in artikli_pretrage:
                if artikal["sifra"]:
                    worksheet.cell(row=row_idx, column=2).value = artikal["sifra"]
                else:
                    worksheet.cell(row=row_idx, column=2).value = '---'
                if artikal["OEbroj"]:
                    worksheet.cell(row=row_idx, column=3).value = artikal["OEbroj"]
                else:
                    worksheet.cell(row=row_idx, column=3).value = '---'
                if artikal["opis"]:
                    worksheet.cell(row=row_idx, column=4).value = artikal["opis"]
                else:
                    worksheet.cell(row=row_idx, column=4).value = '---'
                if artikal["brend"]:
                    worksheet.cell(row=row_idx, column=5).value = artikal["brend"]
                else:
                    worksheet.cell(row=row_idx, column=5).value = '---'
                if artikal["cena"]:
                    worksheet.cell(row=row_idx, column=6).value = artikal["cena"]
                else:
                    worksheet.cell(row=row_idx, column=6).value = '---'
                if len(artikal["stanje"]) > 0:
                    if artikal["stanje"][0] == "GRESKA":
                        worksheet.cell(row=row_idx, column=7).value = 'GRESKA'
                    else:
                        worksheet.cell(row=row_idx, column=7).value = 'IMA'
                else:
                    worksheet.cell(row=row_idx, column=7).value = 'NEMA'
                if CUVAJ_SVE:
                    for col_num, lok in enumerate(artikal["stanje"]):
                        worksheet.cell(row=row_idx, column=8 + col_num).value = lok
                row_idx += 1

            book.save(XLSX_FILE_TEMP)
            if os.path.exists(XLSX_FILE):
                os.remove(XLSX_FILE)
            shutil.copyfile(XLSX_FILE_TEMP, XLSX_FILE)
        if _artikal_ != artikli_za_pretragu[-1]:
            print "\nCekam {sekunde} sekundi".format(sekunde=STRING_PAUZA)
            time.sleep(PAUZA_IZMEDJU_PRETRAGA)
    driver.quit()
    print "\nUspesna pretraga!"


if rvtrade:
    # get the path of ChromeDriverServer
    dir = os.path.dirname(__file__)
    chrome_driver_path = "chromedriver.exe"

    # create a new Chrome session
    driver = webdriver.Chrome(chrome_driver_path)
    driver.implicitly_wait(30)
    # mora full screen , u suprotnom nece naci pretragu !!
    driver.maximize_window()

    driver.get("http://rvtrade.atit-solutions.eu/admin/login_login.php")
    username = driver.find_element_by_name("email")
    password = driver.find_element_by_name("password")

    username.send_keys(korisnik_rvtrade)
    password.send_keys(lozinka_rvtrade)

    driver.find_element_by_class_name("click_button").click()


    time.sleep(2)

    for _artikal_ in artikli_za_pretragu:
        print "Pretrazujem artikal '{0}'...\n".format(_artikal_)
        artikli_pretrage = []
        searchbox = driver.find_element_by_id("search_pattern")
        searchbox.clear()
        # unesi artikal za pretragu
        searchbox.send_keys(_artikal_)
        time.sleep(1)

        driver.find_element_by_id("search_button_1").click()
        # ucitano = driver.execute_script("return document.readyState")
        loader = driver.find_element_by_css_selector(".ajax_loader.search_articles_loader")
        visible = loader.is_displayed()
        # cekaj dok ucitava tabelu
        prolaz = 0
        while visible:
            time.sleep(1)
            visible = loader.is_displayed()
            prolaz += 1
            if prolaz > 60:
                visible = False

        # tabela = driver.find_element_by_id("search_articles_table")
        html = driver.page_source
        # ucitavanje cele stranice
        soup = BeautifulSoup(html, "lxml")
        uspesna_pretraga = True

        tabela = soup.find('table', id="search_articles_table")
        ucitano = driver.execute_script("return document.readyState")
        if "Nema rezultata za pretragu" not in tabela.text:
            head = tabela.find('thead')
            body = tabela.find('tbody')
            svi_artikli = body.find_all('tr')
            for a in svi_artikli:

                podaci = a.find_all('td')

                artikal_recnik = {}
                brojevi = podaci[1].find_all('nobr')
                broj_1 = ""
                if len(brojevi) > 1:
                    broj_1 = brojevi[1].text
                broj_2 = brojevi[0].text
                marka = podaci[2].text
                opis = podaci[4].text
                opis_2 = podaci[4].find('span')
                if opis_2 != None:
                    opis_2 = opis_2.text
                    opis = opis.replace(opis_2, "")
                else:
                    opis_2 = ""
                kolicina = podaci[6].text
                if int(kolicina) > 0:
                    stanje = 'IMA'
                else:
                    stanje = 'NEMA'
                cena = podaci[9].text
                cena = cena.replace(",", ".")
                if cena != "":
                    cena = float(cena)
                artikal_recnik['marka'] = marka
                artikal_recnik['opis'] = opis
                artikal_recnik['opis_2'] = opis_2
                artikal_recnik['stanje'] = stanje
                artikal_recnik['cena'] = cena
                artikal_recnik['kolicina'] = kolicina
                artikal_recnik['broj_1'] = broj_1
                artikal_recnik['broj_2'] = broj_2
                artikli_pretrage.append(artikal_recnik)

        worksheet = book['rvtrade']
        row_idx = worksheet.max_row + 2
        worksheet.cell(row=row_idx, column=1).value = _artikal_
        if artikli_pretrage != []:
            worksheet.cell(row=row_idx + 1, column=2).value = "SIFRA"
            worksheet.cell(row=row_idx + 1, column=3).value = "SIFRA_2"
            worksheet.cell(row=row_idx + 1, column=4).value = "MARKA"
            worksheet.cell(row=row_idx + 1, column=5).value = "OPIS"
            worksheet.cell(row=row_idx + 1, column=6).value = "OPIS_2"
            worksheet.cell(row=row_idx + 1, column=7).value = "STANJE"
            worksheet.cell(row=row_idx + 1, column=8).value = "KOLICINA"
            worksheet.cell(row=row_idx + 1, column=9).value = "CENA"
            row_idx += 2
            for artikal in artikli_pretrage:
                worksheet.cell(row=row_idx, column=2).value = artikal["broj_1"]
                worksheet.cell(row=row_idx, column=3).value = artikal["broj_2"]
                worksheet.cell(row=row_idx, column=4).value = artikal["marka"]
                worksheet.cell(row=row_idx, column=5).value = artikal["opis"]
                worksheet.cell(row=row_idx, column=6).value = artikal["opis_2"]
                worksheet.cell(row=row_idx, column=7).value = artikal["stanje"]
                worksheet.cell(row=row_idx, column=8).value = artikal["kolicina"]
                worksheet.cell(row=row_idx, column=9).value = artikal["cena"]

                row_idx += 1

        else:
            worksheet.cell(row=row_idx + 1, column=2).value = "Artikal ne postoji u bazi"

            row_idx += 1

        book.save(XLSX_FILE_TEMP)
        if os.path.exists(XLSX_FILE):
            os.remove(XLSX_FILE)
        shutil.copyfile(XLSX_FILE_TEMP, XLSX_FILE)
        if _artikal_ != artikli_za_pretragu[-1]:
            print "\nCekam {sekunde} sekundi".format(sekunde=STRING_PAUZA)
            time.sleep(PAUZA_IZMEDJU_PRETRAGA)

    driver.quit()
    print "\nUspesna pretraga!"